"""控排企业履约策略：CRUD、配置、策略运行门面（JSON 文件存储版）。

函数签名与参考项目 pdf_trans 的 carbon_compliance_service.py 保持一致，
`db` 参数保留但接受 None（无 SQLAlchemy 会话），数据存储于
backend/data/carbon_compliance.json。
"""

from __future__ import annotations

import io
import json
import os
import uuid
from datetime import datetime, timezone
from typing import Any

from app.services.carbon_compliance.accounting import AccountingInput, compute_accounting
from app.services.carbon_compliance.alerts import build_alerts
from app.services.carbon_compliance.carry_forward import compute_carry_forward
from app.services.carbon_compliance.compliance import eligible_ccer_qty
from app.services.carbon_compliance.defaults import (
    INDUSTRIES,
    INDUSTRY_LABELS,
    RISK_PROFILES,
    deep_merge,
    default_settings,
)
from app.services.carbon_compliance.market_cycle import judge_market_cycle
from app.services.carbon_compliance.report_export import strategy_run_to_markdown
from app.services.carbon_compliance.strategy_engine import build_strategy_payload

try:
    from openpyxl import Workbook, load_workbook
except ImportError:  # pragma: no cover
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

_DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "..", "data")
_DATA_PATH = os.path.join(_DATA_DIR, "carbon_compliance.json")

# 平台无用户系统，固定默认用户
DEFAULT_USER_ID = "default"


def _now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def _empty_store() -> dict:
    return {
        "settings": {},
        "enterprises": [],
        "market_cea": [],
        "market_ccer": [],
        "market_energy": [],
    }


def _load_store() -> dict:
    if not os.path.exists(_DATA_PATH):
        store = _empty_store()
        _save_store(store)
        return store
    try:
        with open(_DATA_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
    except Exception:
        data = _empty_store()
    return {
        "settings": data.get("settings") or {},
        "enterprises": data.get("enterprises") or [],
        "market_cea": data.get("market_cea") or [],
        "market_ccer": data.get("market_ccer") or [],
        "market_energy": data.get("market_energy") or [],
    }


def _save_store(store: dict) -> None:
    os.makedirs(_DATA_DIR, exist_ok=True)
    tmp = _DATA_PATH + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(store, f, ensure_ascii=False, indent=2)
    os.replace(tmp, _DATA_PATH)


class _Row(dict):
    """支持属性访问的 dict（兼容流水线 ORM 属性访问方式）。

    缺失键返回 None（对齐参考版 ORM 列默认值语义，避免属性访问抛错）。
    """

    def __getattr__(self, key: str):
        return self.get(key)


def _uuid() -> str:
    return str(uuid.uuid4())


def _find_enterprise(store: dict, enterprise_id: str) -> dict | None:
    for e in store.get("enterprises") or []:
        if str(e.get("id")) == str(enterprise_id):
            return e
    return None


def get_settings(db: Any = None) -> dict:
    store = _load_store()
    return deep_merge(default_settings(), store.get("settings") or {})


def update_settings(db: Any = None, payload: dict | None = None) -> dict:
    store = _load_store()
    merged = deep_merge(default_settings(), deep_merge(store.get("settings") or {}, payload or {}))
    store["settings"] = merged
    _save_store(store)
    return deep_merge(default_settings(), store["settings"])


def list_enterprises(db: Any = None, user_id: str = DEFAULT_USER_ID) -> list:
    store = _load_store()
    rows = [e for e in store.get("enterprises") or [] if str(e.get("user_id")) == str(user_id)]
    rows.sort(key=lambda e: e.get("updated_at") or "", reverse=True)
    return [_Row(e) for e in rows]


def get_enterprise(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = ""
) -> _Row | None:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent or str(ent.get("user_id")) != str(user_id):
        return None
    return _Row(ent)


def create_enterprise(db: Any = None, user_id: str = DEFAULT_USER_ID, data: dict | None = None) -> _Row:
    data = data or {}
    industry = data.get("industry") or ""
    if industry not in INDUSTRIES:
        raise ValueError(f"industry must be one of {INDUSTRIES}")
    risk = data.get("risk_profile") or "balanced"
    if risk not in RISK_PROFILES:
        raise ValueError(f"risk_profile must be one of {RISK_PROFILES}")
    raw_name = str(data.get("name") or "").strip()
    if not raw_name:
        label = INDUSTRY_LABELS.get(industry, industry)
        uscc = str(data.get("uscc") or "").strip()
        raw_name = f"{label}履约主体" + (f"-{uscc[-6:]}" if uscc else "")
    now = _now_iso()
    ent = {
        "id": _uuid(),
        "user_id": str(user_id),
        "name": raw_name,
        "uscc": str(data.get("uscc") or "").strip(),
        "industry": industry,
        "market_start_year": int(data["market_start_year"]),
        "compliance_cycle": str(data.get("compliance_cycle") or "annual"),
        "risk_profile": risk,
        "annual_budget_cap": float(data.get("annual_budget_cap") or 0),
        "single_trade_limit": float(data.get("single_trade_limit") or 0),
        "enterprise_attrs": dict(data.get("enterprise_attrs") or {}),
        "created_at": now,
        "updated_at": now,
        "emissions": [],
        "forecasts": [],
        "cea_holdings": [],
        "cea_trades": [],
        "ccer_holdings": [],
        "ccer_trades": [],
        "green_power": [],
        "green_certs": [],
        "strategy_runs": [],
        "alerts": [],
    }
    store = _load_store()
    store["enterprises"].append(ent)
    _save_store(store)
    return _Row(ent)


def update_enterprise(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    data = data or {}
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent or str(ent.get("user_id")) != str(user_id):
        raise LookupError("enterprise not found")
    if "name" in data and data["name"] is not None:
        ent["name"] = str(data["name"]).strip()
    if "uscc" in data:
        ent["uscc"] = str(data.get("uscc") or "").strip()
    if "industry" in data and data["industry"] is not None:
        if data["industry"] not in INDUSTRIES:
            raise ValueError(f"industry must be one of {INDUSTRIES}")
        ent["industry"] = data["industry"]
    if "market_start_year" in data and data["market_start_year"] is not None:
        ent["market_start_year"] = int(data["market_start_year"])
    if "compliance_cycle" in data and data["compliance_cycle"] is not None:
        ent["compliance_cycle"] = str(data["compliance_cycle"])
    if "risk_profile" in data and data["risk_profile"] is not None:
        if data["risk_profile"] not in RISK_PROFILES:
            raise ValueError(f"risk_profile must be one of {RISK_PROFILES}")
        ent["risk_profile"] = data["risk_profile"]
    if "annual_budget_cap" in data and data["annual_budget_cap"] is not None:
        ent["annual_budget_cap"] = float(data["annual_budget_cap"])
    if "single_trade_limit" in data and data["single_trade_limit"] is not None:
        ent["single_trade_limit"] = float(data["single_trade_limit"])
    if "enterprise_attrs" in data and data["enterprise_attrs"] is not None:
        ent["enterprise_attrs"] = dict(data["enterprise_attrs"])
    ent["updated_at"] = _now_iso()
    _save_store(store)
    return _Row(ent)


def delete_enterprise(db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "") -> None:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent or str(ent.get("user_id")) != str(user_id):
        raise LookupError("enterprise not found")
    store["enterprises"] = [e for e in store["enterprises"] if str(e.get("id")) != str(enterprise_id)]
    _save_store(store)


def _upsert_emission_year(store: dict, enterprise_id: str, data: dict) -> _Row:
    ent = _find_enterprise(store, enterprise_id)
    year = int(data["year"])
    row = None
    for r in ent.get("emissions") or []:
        if int(r.get("year")) == year:
            row = r
            break
    if not row:
        row = {"id": _uuid(), "year": year}
        ent.setdefault("emissions", []).append(row)
    for field in (
        "verified_total",
        "scope1_combustion",
        "scope1_process",
        "scope2_power",
        "purchased_mwh",
        "historical_gap",
        "ccer_used",
    ):
        if field in data:
            row[field] = data[field]
    if "monthly_detail" in data:
        row["monthly_detail"] = dict(data["monthly_detail"] or {})
    return _Row(row)


def list_emission_years(db: Any = None, enterprise_id: str = "") -> list:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return []
    rows = sorted(ent.get("emissions") or [], key=lambda r: r.get("year"), reverse=True)
    return [_Row(r) for r in rows]


def upsert_emission_year(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    row = _upsert_emission_year(store, enterprise_id, data or {})
    _save_store(store)
    return row


def delete_emission_year(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", year: int = 0
) -> None:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    before = len(ent.get("emissions") or [])
    ent["emissions"] = [r for r in ent.get("emissions") or [] if int(r.get("year")) != int(year)]
    if len(ent.get("emissions")) == before:
        raise LookupError("emission year not found")
    _save_store(store)


def upsert_forecast(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    year = int(data["year"])
    row = None
    for r in ent.get("forecasts") or []:
        if int(r.get("year")) == year:
            row = r
            break
    if not row:
        row = {"id": _uuid(), "year": year}
        ent.setdefault("forecasts", []).append(row)
    row["forecast_total"] = float(data.get("forecast_total") or 0)
    row["capacity_plan"] = str(data.get("capacity_plan") or "")
    row["abatement_projects"] = list(data.get("abatement_projects") or [])
    row["production_plan"] = str(data.get("production_plan") or "")
    _save_store(store)
    return _Row(row)


def list_forecasts(db: Any = None, enterprise_id: str = "") -> list:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return []
    rows = sorted(ent.get("forecasts") or [], key=lambda r: r.get("year"), reverse=True)
    return [_Row(r) for r in rows]


def upsert_cea_holding(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    vy = int(data["vintage_year"])
    row = None
    for r in ent.get("cea_holdings") or []:
        if int(r.get("vintage_year")) == vy:
            row = r
            break
    if not row:
        row = {"id": _uuid(), "vintage_year": vy}
        ent.setdefault("cea_holdings", []).append(row)
    for field in (
        "free_quota",
        "carry_forward_qty",
        "net_sell_qty",
        "avg_cost",
        "estimated_free_quota",
        "sellable_cap",
    ):
        if field in data:
            row[field] = data[field]
    # 兼容旧字段名
    if "carry_forward_qty" not in data and "expired_qty" in data:
        row["carry_forward_qty"] = float(data.get("expired_qty") or 0)
    _save_store(store)
    return _Row(row)


def list_cea_holdings(db: Any = None, enterprise_id: str = "") -> list:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return []
    rows = sorted(ent.get("cea_holdings") or [], key=lambda r: r.get("vintage_year"), reverse=True)
    return [_Row(r) for r in rows]


def delete_cea_holding(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", vintage_year: int = 0
) -> None:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    before = len(ent.get("cea_holdings") or [])
    ent["cea_holdings"] = [
        r for r in ent.get("cea_holdings") or [] if int(r.get("vintage_year")) != int(vintage_year)
    ]
    if len(ent.get("cea_holdings")) == before:
        raise LookupError("cea holding not found")
    _save_store(store)


def get_cea_holding_for_year(
    db: Any = None, enterprise_id: str = "", vintage_year: int = 0
) -> _Row | None:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return None
    for r in ent.get("cea_holdings") or []:
        if int(r.get("vintage_year")) == int(vintage_year):
            return _Row(r)
    return None


def add_cea_trade(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    side = data.get("side")
    if side not in ("buy", "sell"):
        raise ValueError("side must be buy or sell")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    row = {
        "id": _uuid(),
        "enterprise_id": enterprise_id,
        "side": side,
        "qty": float(data["qty"]),
        "price": float(data.get("price") or 0),
        "note": str(data.get("note") or ""),
        "traded_at": _now_iso(),
    }
    ent.setdefault("cea_trades", []).append(row)
    _save_store(store)
    return _Row(row)


def list_cea_trades(db: Any = None, enterprise_id: str = "") -> list:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return []
    rows = sorted(ent.get("cea_trades") or [], key=lambda r: r.get("traded_at"), reverse=True)
    return [_Row(r) for r in rows]


def cea_net_sell_qty(db: Any = None, enterprise_id: str = "") -> float:
    """从交易流水汇总净卖出 = 卖出量 − 买入量（万吨）。"""
    sold = 0.0
    bought = 0.0
    for t in list_cea_trades(db, enterprise_id):
        q = float(t.get("qty") or 0)
        if t.get("side") == "sell":
            sold += q
        elif t.get("side") == "buy":
            bought += q
    return sold - bought


def resolve_cea_net_sell(
    db: Any = None, enterprise_id: str = "", compliance_year: int = 0
) -> float:
    """结转测算用净卖出：优先取履约年 CEA 台账「当前净卖出」，否则回退交易流水汇总。"""
    row = get_cea_holding_for_year(db, enterprise_id, compliance_year)
    if row is not None:
        return float(row.get("net_sell_qty") or 0)
    return cea_net_sell_qty(db, enterprise_id)


def upsert_ccer_holding(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    data = data or {}
    expire_at = data.get("expire_at")
    if isinstance(expire_at, str) and expire_at:
        expire_at = expire_at[:10]
    qty = max(0.0, float(data.get("qty") or 0))
    if data.get("eligible_qty") is not None:
        eligible = max(0.0, float(data["eligible_qty"]))
        if eligible <= 0 and qty > 0:
            eligible = qty
        elif qty > 0:
            eligible = min(eligible, qty)
    else:
        eligible = qty
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    holding_id = str(data.get("id") or _uuid())
    row = None
    for r in ent.get("ccer_holdings") or []:
        if str(r.get("id")) == holding_id:
            row = r
            break
    if not row:
        row = {"id": holding_id}
        ent.setdefault("ccer_holdings", []).append(row)
    row["enterprise_id"] = enterprise_id
    row["project_type"] = str(data.get("project_type") or "")
    row["issue_year"] = int(data["issue_year"])
    row["expire_at"] = expire_at
    row["qty"] = qty
    row["cost"] = float(data.get("cost") or 0)
    row["eligible_qty"] = eligible
    row["linked_green_cert"] = bool(data.get("linked_green_cert") or False)
    _save_store(store)
    return _Row(row)


def list_ccer_holdings(db: Any = None, enterprise_id: str = "") -> list:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return []
    rows = sorted(ent.get("ccer_holdings") or [], key=lambda r: r.get("issue_year"), reverse=True)
    return [_Row(r) for r in rows]


def delete_ccer_holding(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", holding_id: str = ""
) -> None:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    before = len(ent.get("ccer_holdings") or [])
    ent["ccer_holdings"] = [
        r for r in ent.get("ccer_holdings") or [] if str(r.get("id")) != str(holding_id)
    ]
    if len(ent.get("ccer_holdings")) == before:
        raise LookupError("ccer holding not found")
    _save_store(store)


def add_ccer_trade(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    side = data.get("side")
    if side not in ("buy", "sell"):
        raise ValueError("side must be buy or sell")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    row = {
        "id": _uuid(),
        "enterprise_id": enterprise_id,
        "side": side,
        "qty": float(data["qty"]),
        "price": float(data.get("price") or 0),
        "note": str(data.get("note") or ""),
        "traded_at": _now_iso(),
    }
    ent.setdefault("ccer_trades", []).append(row)
    _save_store(store)
    return _Row(row)


def upsert_green_power(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    year = int(data["year"])
    row = None
    for r in ent.get("green_power") or []:
        if int(r.get("year")) == year:
            row = r
            break
    if not row:
        row = {"id": _uuid(), "year": year}
        ent.setdefault("green_power", []).append(row)
    row["market_green_mwh"] = float(data.get("market_green_mwh") or 0)
    row["self_gen_mwh"] = float(data.get("self_gen_mwh") or 0)
    row["premium_per_mwh"] = float(data.get("premium_per_mwh") or 0)
    row["contract_ref"] = str(data.get("contract_ref") or "")
    _save_store(store)
    return _Row(row)


def list_green_power(db: Any = None, enterprise_id: str = "") -> list:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return []
    rows = sorted(ent.get("green_power") or [], key=lambda r: r.get("year"), reverse=True)
    return [_Row(r) for r in rows]


def upsert_green_cert(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", data: dict | None = None
) -> _Row:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    data = data or {}
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    cert_id = str(data.get("id") or _uuid())
    row = None
    for r in ent.get("green_certs") or []:
        if str(r.get("id")) == cert_id:
            row = r
            break
    if not row:
        row = {"id": cert_id}
        ent.setdefault("green_certs", []).append(row)
    row["enterprise_id"] = enterprise_id
    row["year"] = int(data["year"])
    row["qty"] = float(data.get("qty") or 0)
    row["unit_price"] = float(data.get("unit_price") or 0)
    row["retired"] = bool(data.get("retired") or False)
    row["ren_weight_target"] = data.get("ren_weight_target")
    _save_store(store)
    return _Row(row)


def list_green_certs(db: Any = None, enterprise_id: str = "") -> list:
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    if not ent:
        return []
    rows = sorted(ent.get("green_certs") or [], key=lambda r: r.get("year"), reverse=True)
    return [_Row(r) for r in rows]


def upsert_market_cea(db: Any = None, data: dict | None = None) -> _Row:
    data = data or {}
    ym = str(data["year_month"])
    store = _load_store()
    row = None
    for r in store.get("market_cea") or []:
        if str(r.get("year_month")) == ym:
            row = r
            break
    if not row:
        row = {"id": _uuid(), "year_month": ym}
        store.setdefault("market_cea", []).append(row)
    row["avg_price"] = float(data["avg_price"])
    row["high"] = data.get("high")
    row["low"] = data.get("low")
    row["period_tag"] = str(data.get("period_tag") or "")
    _save_store(store)
    return _Row(row)


def list_market_cea(db: Any = None, limit: int = 60) -> list:
    store = _load_store()
    rows = sorted(store.get("market_cea") or [], key=lambda r: r.get("year_month"), reverse=True)
    return [_Row(r) for r in rows[:limit]]


def upsert_market_ccer(db: Any = None, data: dict | None = None) -> _Row:
    data = data or {}
    ym = str(data["year_month"])
    ptype = str(data.get("project_type") or "general")
    store = _load_store()
    row = None
    for r in store.get("market_ccer") or []:
        if str(r.get("year_month")) == ym and str(r.get("project_type") or "general") == ptype:
            row = r
            break
    if not row:
        row = {"id": _uuid(), "year_month": ym, "project_type": ptype}
        store.setdefault("market_ccer", []).append(row)
    row["avg_price"] = float(data["avg_price"])
    _save_store(store)
    return _Row(row)


def list_market_ccer(db: Any = None, limit: int = 60) -> list:
    store = _load_store()
    rows = sorted(store.get("market_ccer") or [], key=lambda r: r.get("year_month"), reverse=True)
    return [_Row(r) for r in rows[:limit]]


def upsert_market_energy(db: Any = None, data: dict | None = None) -> _Row:
    data = data or {}
    ym = str(data["year_month"])
    region = str(data.get("region") or "national")
    store = _load_store()
    row = None
    for r in store.get("market_energy") or []:
        if str(r.get("year_month")) == ym and str(r.get("region") or "national") == region:
            row = r
            break
    if not row:
        row = {"id": _uuid(), "year_month": ym, "region": region}
        store.setdefault("market_energy", []).append(row)
    row["green_premium"] = data.get("green_premium")
    row["grec_price"] = data.get("grec_price")
    row["coal_price"] = data.get("coal_price")
    _save_store(store)
    return _Row(row)


def list_market_energy(db: Any = None, limit: int = 60) -> list:
    store = _load_store()
    rows = sorted(store.get("market_energy") or [], key=lambda r: r.get("year_month"), reverse=True)
    return [_Row(r) for r in rows[:limit]]


async def sync_market_quotes(db: Any = None) -> dict[str, Any]:
    """拉取外部 CEA/CCER 行情并写入月度表（复用目标项目同步行情函数）。"""
    from app.services.carbon_compliance.market_sync import (
        fetch_cneeex_daily_quotes_sync,
    )

    errors: list[str] = []
    written: dict[str, Any] = {"cea": None, "ccer": None, "cea_months": 0}
    try:
        hist = fetch_cneeex_daily_quotes_sync() or []
        by_month: dict[str, list[float]] = {}
        for q in hist:
            t = str(q.get("t") or "")
            ym = t[:7]
            by_month.setdefault(ym, []).append(float(q.get("close") or q.get("price") or 0))
        months = sorted(by_month.keys())
        for ym in months:
            prices = by_month[ym]
            upsert_market_cea(
                None,
                {
                    "year_month": ym,
                    "avg_price": sum(prices) / len(prices),
                    "high": max(prices),
                    "low": min(prices),
                    "period_tag": _period_tag(int(ym.split("-")[1])),
                },
            )
        written["cea_months"] = len(months)
        if months:
            last_ym = months[-1]
            row = None
            for r in list_market_cea(None, limit=1):
                row = r
            written["cea"] = {
                "year_month": last_ym,
                "avg_price": row.get("avg_price") if row else None,
                "high": row.get("high") if row else None,
                "low": row.get("low") if row else None,
            }
    except Exception as exc:  # noqa: BLE001
        errors.append(f"cea: {exc}")

    settings = get_settings(db)
    integrations = dict(settings.get("integrations") or {})
    integrations["last_sync_at"] = _now_iso()
    integrations["last_sync_ok"] = bool(written["cea"] or written["ccer"])
    integrations["last_sync_detail"] = {"written": written, "errors": errors}
    update_settings(db, {"integrations": integrations})

    return {
        "ok": bool(written["cea"] or written["ccer"]),
        "written": written,
        "errors": errors,
    }


def _period_tag(month: int) -> str:
    if month <= 2:
        return "Q1"
    if month <= 5:
        return "Q2"
    if month <= 8:
        return "Q3"
    return "Q4"


def should_auto_sync_market(settings: dict | None = None) -> bool:
    """根据 market_sync_period 与 last_sync_at 判断是否应自动同步。"""
    cfg = (settings or {}).get("integrations") or {}
    if cfg.get("market_sync_enabled") is False:
        return False
    period = str(cfg.get("market_sync_period") or "day").lower()
    last = cfg.get("last_sync_at")
    if not last:
        return True
    try:
        last_dt = datetime.fromisoformat(str(last).replace("Z", "+00:00"))
    except Exception:
        return True
    now = datetime.now(timezone.utc)
    if last_dt.tzinfo is None:
        last_dt = last_dt.replace(tzinfo=timezone.utc)
    age_h = (now - last_dt.astimezone(timezone.utc)).total_seconds() / 3600.0
    if period in ("hour", "hourly"):
        return age_h >= 1.0
    if period in ("day", "daily"):
        return age_h >= 12.0
    return age_h >= 24.0


def _holding_to_dict(h: _Row) -> dict:
    return {
        "eligible_qty": h.get("eligible_qty"),
        "qty": h.get("qty"),
        "expire_at": h.get("expire_at"),
        "linked_green_cert": h.get("linked_green_cert"),
    }


def run_strategy(
    db: Any = None,
    user_id: str = DEFAULT_USER_ID,
    enterprise_id: str = "",
    compliance_year: int = 0,
    *,
    notify: bool = True,
) -> _Row:
    ent = get_enterprise(db, user_id, enterprise_id)
    if not ent:
        raise LookupError("enterprise not found")
    settings = get_settings(db)
    industry_params = (settings.get("industry_params") or {}).get(ent.get("industry")) or {}
    cost = settings.get("cost") or {}
    compliance_cfg = settings.get("compliance") or {}

    emission = None
    for r in list_emission_years(db, enterprise_id):
        if int(r.get("year")) == int(compliance_year):
            emission = r
            break
    forecast = None
    for r in list_forecasts(db, enterprise_id):
        if int(r.get("year")) == int(compliance_year):
            forecast = r
            break

    cea_rows = list_cea_holdings(db, enterprise_id)
    free_cea = 0.0
    sellable = 0.0
    allocated_free = 0.0  # 当年新分配（不含结转）
    carry_in = 0.0
    market_start_year = int(ent.get("market_start_year") or 0)
    for c in cea_rows:
        if int(c.get("vintage_year") or 0) == int(compliance_year) or int(
            c.get("vintage_year") or 0
        ) >= market_start_year:
            allocated = max(0.0, float(c.get("free_quota") or 0))
            carried = max(0.0, float(c.get("carry_forward_qty") or 0))
            usable = allocated + carried
            free_cea += usable
            if int(c.get("vintage_year") or 0) == int(compliance_year):
                allocated_free += allocated
                carry_in += carried
            if c.get("estimated_free_quota") and int(c.get("vintage_year") or 0) == int(
                compliance_year
            ):
                free_cea = max(free_cea, float(c.get("estimated_free_quota")) + carried)
            if c.get("sellable_cap") is not None:
                sellable += float(c.get("sellable_cap"))
            else:
                sellable += usable

    ccer_rows = list_ccer_holdings(db, enterprise_id)
    own_ccer = eligible_ccer_qty([_holding_to_dict(h) for h in ccer_rows])

    grid_factor = float(
        industry_params.get("grid_emission_factor")
        or cost.get("grid_emission_factor")
        or 0.5703
    )
    ccer_ratio = float(compliance_cfg.get("ccer_max_ratio") or 0.05)

    scope1_c = float(emission.get("scope1_combustion") or 0.0) if emission else 0.0
    scope1_p = float(emission.get("scope1_process") or 0.0) if emission else 0.0
    scope2 = float(emission.get("scope2_power") or 0.0) if emission else 0.0
    purchased = float(emission.get("purchased_mwh") or 0.0) if emission else 0.0

    verified_override = None
    if emission and emission.get("verified_total") is not None:
        verified_override = float(emission.get("verified_total"))
    elif not emission and forecast and forecast.get("forecast_total"):
        verified_override = float(forecast.get("forecast_total"))

    acc = compute_accounting(
        AccountingInput(
            scope1_combustion=scope1_c,
            scope1_process=scope1_p,
            scope2_power=scope2,
            purchased_mwh=purchased,
            market_green_mwh=0.0,
            self_gen_mwh=0.0,
            free_cea_quota=free_cea,
            own_ccer_eligible=own_ccer,
            grid_emission_factor=grid_factor,
            ccer_max_ratio=ccer_ratio,
            verified_override=verified_override,
        )
    )

    cea_market = list_market_cea(db, limit=60)
    prices = [float(r.get("avg_price")) for r in reversed(cea_market)]
    current_cea = prices[-1] if prices else 80.0
    ccer_market = list_market_ccer(db, limit=24)
    current_ccer = float(ccer_market[0].get("avg_price")) if ccer_market else 60.0

    market = judge_market_cycle(
        prices,
        current_cea,
        current_ccer,
        low_percentile=float(compliance_cfg.get("price_low_percentile") or 0.30),
        mid_percentile=float(compliance_cfg.get("price_mid_percentile") or 0.70),
    )

    # 日度碳价预测：用「下一交易日/近期」预测价作为当日无挂单时的锚定
    price_forecast_summary = None
    cea_predicted = None
    ccer_predicted = None
    try:
        from app.services.carbon_compliance.market_sync import fetch_cneeex_daily_quotes_sync
        from app.services.carbon_compliance.price_forecast import forecast_cea_to_year_end

        hist = fetch_cneeex_daily_quotes_sync()
        if not hist:
            hist = [
                {"t": f"{r.get('year_month')}-15", "close": float(r.get("avg_price"))}
                for r in reversed(cea_market)
            ]
        fc = forecast_cea_to_year_end(hist)
        if fc.get("ok"):
            price_forecast_summary = fc.get("summary")
            pts = fc.get("points") or []
            if pts:
                cea_predicted = float(pts[0].get("price") or pts[0].get("close") or current_cea)
            ye = float((price_forecast_summary or {}).get("year_end_price") or current_cea)
            if acc.compliance_gap > 0 and ye > current_cea * 1.03:
                market.rationale = (
                    (market.rationale or "")
                    + f"；日度预测年底约 {ye:.1f} 元/吨（高于现价），宜评估提前分批采购"
                )
            elif acc.compliance_gap < 0 and ye > current_cea * 1.03:
                market.rationale = (
                    (market.rationale or "")
                    + f"；日度预测年底约 {ye:.1f} 元/吨，盈余配额可择高位窗口出售"
                )
    except Exception:  # noqa: BLE001
        price_forecast_summary = None

    # CCER 无公开分时：用最新日均作当日预测锚（可后续接独立预测）
    ccer_predicted = current_ccer

    # 结转测算：年末持仓≈清缴后盈余；基础默认取当年免费配额；
    # 净卖出优先取履约年 CEA 台账「当前净卖出」
    net_sell = resolve_cea_net_sell(db, enterprise_id, compliance_year)
    year_end_holding = max(0.0, -min(0.0, float(acc.compliance_gap)))
    carry_base = float(compliance_cfg.get("carry_base_qty") or 0)
    if carry_base <= 0:
        carry_base = allocated_free if allocated_free > 0 else max(0.0, float(acc.free_cea_quota))
    carry = compute_carry_forward(
        base_qty=carry_base,
        net_sell=net_sell,
        year_end_holding=year_end_holding,
        net_sell_multiplier=float(compliance_cfg.get("carry_net_sell_multiplier") or 1.5),
        deadline_md=str(compliance_cfg.get("carry_forward_deadline_md") or "06-10"),
        deadline_year=compliance_year + 1,  # 履约年后的结转日（如 2025 履约 → 2026-06-10）
    )

    payload = build_strategy_payload(
        risk_profile=ent.get("risk_profile"),
        annual_budget_cap=ent.get("annual_budget_cap"),
        single_trade_limit=ent.get("single_trade_limit"),
        settings=settings,
        accounting=acc,
        market=market,
        cea_price=current_cea,
        ccer_price=current_ccer,
        sellable_cea=sellable,
        cea_price_predicted=cea_predicted,
        ccer_price_predicted=ccer_predicted,
        carry_forward=carry,
    )
    if price_forecast_summary:
        payload["price_forecast"] = price_forecast_summary
        tags = dict(payload.get("market_tags") or {})
        tags["price_forecast"] = price_forecast_summary
        payload["market_tags"] = tags
    payload["carry_forward"] = carry.to_dict()
    payload["accounting"] = {
        **payload["accounting"],
        "allocated_free_cea": allocated_free,
        "carry_forward_qty": carry_in,
        "usable_cea_total": float(acc.free_cea_quota),
    }
    tags = dict(payload.get("market_tags") or {})
    tags["carry_forward"] = carry.to_dict()
    payload["market_tags"] = tags
    report_md = strategy_run_to_markdown(
        enterprise_name=ent.get("name"),
        compliance_year=compliance_year,
        accounting=payload["accounting"],
        market_tags=payload["market_tags"],
        plans=payload["plans"],
    )

    store = _load_store()
    ent_row = _find_enterprise(store, enterprise_id)
    run = {
        "id": _uuid(),
        "enterprise_id": enterprise_id,
        "user_id": str(user_id),
        "compliance_year": int(compliance_year),
        "accounting_snapshot": payload["accounting"],
        "market_tags": payload["market_tags"],
        "plans": payload["plans"],
        "status": "completed",
        "report_md": report_md,
        "created_at": _now_iso(),
    }
    ent_row.setdefault("strategy_runs", []).append(run)

    ccer_planned = 0.0
    for p in payload["plans"]:
        for a in p.get("actions") or []:
            if a.get("action") in ("use_ccer", "buy_ccer"):
                ccer_planned = max(ccer_planned, float(a.get("qty") or 0))
                break

    alert_dicts = build_alerts(
        compliance_year=compliance_year,
        clearance_deadline_md=str(industry_params.get("clearance_deadline_md") or "12-31"),
        warn_days=list(compliance_cfg.get("clearance_warn_days") or [90, 30, 15]),
        compliance_gap=acc.compliance_gap,
        ccer_used=ccer_planned,
        ccer_cap=acc.ccer_cap,
        price_band=market.price_band,
        carry_forward=carry,
    )
    for ad in alert_dicts:
        due = ad.get("due_at")
        if isinstance(due, datetime):
            due = due.isoformat()
        ent_row.setdefault("alerts", []).append(
            {
                "id": _uuid(),
                "enterprise_id": enterprise_id,
                "user_id": str(user_id),
                "level": ad["level"],
                "alert_type": ad["alert_type"],
                "message": ad["message"],
                "due_at": due,
                "acked": False,
                "created_at": _now_iso(),
            }
        )
    _save_store(store)

    if notify and alert_dicts:
        try:
            top = alert_dicts[0]
            # 无通知服务时静默跳过
            _ = top
        except Exception:  # noqa: BLE001
            pass

    return _Row(run)


def list_strategy_runs(
    db: Any = None, user_id: str = DEFAULT_USER_ID, enterprise_id: str = "", limit: int = 20
) -> list:
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    store = _load_store()
    ent = _find_enterprise(store, enterprise_id)
    rows = sorted(ent.get("strategy_runs") or [], key=lambda r: r.get("created_at"), reverse=True)
    return [_Row(r) for r in rows[:limit]]


def list_alerts(
    db: Any = None,
    user_id: str = DEFAULT_USER_ID,
    *,
    enterprise_id: str | None = None,
    unacked_only: bool = False,
    limit: int = 50,
) -> list:
    store = _load_store()
    out: list = []
    for e in store.get("enterprises") or []:
        if str(e.get("user_id")) != str(user_id):
            continue
        if enterprise_id and str(e.get("id")) != str(enterprise_id):
            continue
        for a in e.get("alerts") or []:
            if unacked_only and a.get("acked"):
                continue
            out.append(_Row(a))
    out.sort(key=lambda a: a.get("created_at") or "", reverse=True)
    return out[:limit]


def ack_alert(db: Any = None, user_id: str = DEFAULT_USER_ID, alert_id: str = "") -> _Row:
    store = _load_store()
    for e in store.get("enterprises") or []:
        if str(e.get("user_id")) != str(user_id):
            continue
        for a in e.get("alerts") or []:
            if str(a.get("id")) == str(alert_id):
                a["acked"] = True
                _save_store(store)
                return _Row(a)
    raise LookupError("alert not found")


def build_import_template_bytes() -> bytes:
    if Workbook is None:
        raise RuntimeError("openpyxl not installed")
    wb = Workbook()
    ws = wb.active
    ws.title = "emissions"
    ws.append(
        [
            "year",
            "scope1_process_万吨",
            "scope1_combustion_万吨",
            "verified_total_万吨",
        ]
    )
    ws.append([2024, 2, 10, None])
    ws2 = wb.create_sheet("cea")
    ws2.append(
        [
            "vintage_year",
            "free_quota_万吨",
            "carry_forward_qty_万吨",
            "net_sell_qty_万吨",
            "avg_cost",
            "sellable_cap_万吨",
        ]
    )
    ws2.append([2024, 14, 0, 0, 0, 1])
    ws3 = wb.create_sheet("ccer")
    ws3.append(
        [
            "project_type",
            "issue_year",
            "expire_at",
            "qty_万吨",
            "cost",
            "eligible_qty_万吨",
        ]
    )
    ws3.append(["wind", 2023, "2030-12-31", 0.5, 50, 0.5])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def import_enterprise_excel(
    db: Any = None,
    user_id: str = DEFAULT_USER_ID,
    enterprise_id: str = "",
    content: bytes = b"",
) -> dict[str, Any]:
    if load_workbook is None:
        raise RuntimeError("openpyxl not installed")
    if not get_enterprise(db, user_id, enterprise_id):
        raise LookupError("enterprise not found")
    wb = load_workbook(io.BytesIO(content), data_only=True)
    counts = {"emissions": 0, "cea": 0, "ccer": 0}
    wan_fields = frozenset(
        {
            "scope1_process",
            "scope1_combustion",
            "verified_total",
            "free_quota",
            "carry_forward_qty",
            "net_sell_qty",
            "sellable_cap",
            "qty",
            "eligible_qty",
        }
    )

    def _normalize_row(raw: dict[str, Any]) -> dict[str, Any]:
        """Excel 数量列按万吨填写；兼容无后缀的旧表头（仍按万吨解释）。"""
        out: dict[str, Any] = {}
        for key, val in raw.items():
            k = str(key or "").strip()
            if k.endswith("_万吨"):
                k = k[: -len("_万吨")]
            if k == "expired_qty":
                k = "carry_forward_qty"
            if k in wan_fields and val is not None and val != "":
                try:
                    out[k] = float(val)
                except (TypeError, ValueError):
                    out[k] = val
            else:
                out[k] = val
        return out

    if "emissions" in wb.sheetnames:
        ws = wb["emissions"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            data = _normalize_row(
                {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            )
            upsert_emission_year(db, user_id, enterprise_id, data)
            counts["emissions"] += 1

    if "cea" in wb.sheetnames:
        ws = wb["cea"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        for row in rows[1:]:
            if not row or row[0] is None:
                continue
            data = _normalize_row(
                {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            )
            upsert_cea_holding(db, user_id, enterprise_id, data)
            counts["cea"] += 1

    if "ccer" in wb.sheetnames:
        ws = wb["ccer"]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        for row in rows[1:]:
            if not row or all(c is None for c in row):
                continue
            data = _normalize_row(
                {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
            )
            if data.get("expire_at") is not None:
                data["expire_at"] = str(data["expire_at"])[:10]
            upsert_ccer_holding(db, user_id, enterprise_id, data)
            counts["ccer"] += 1

    return counts


def enterprise_to_dict(e: Any) -> dict:
    return {
        "id": str(e.get("id")),
        "name": e.get("name"),
        "uscc": e.get("uscc"),
        "industry": e.get("industry"),
        "market_start_year": e.get("market_start_year"),
        "compliance_cycle": e.get("compliance_cycle"),
        "risk_profile": e.get("risk_profile"),
        "annual_budget_cap": e.get("annual_budget_cap"),
        "single_trade_limit": e.get("single_trade_limit"),
        "enterprise_attrs": e.get("enterprise_attrs") or {},
        "created_at": e.get("created_at"),
        "updated_at": e.get("updated_at"),
    }


def run_to_dict(r: Any) -> dict:
    return {
        "id": str(r.get("id")),
        "enterprise_id": str(r.get("enterprise_id")),
        "compliance_year": r.get("compliance_year"),
        "accounting_snapshot": r.get("accounting_snapshot"),
        "market_tags": r.get("market_tags"),
        "plans": r.get("plans"),
        "status": r.get("status"),
        "report_md": r.get("report_md"),
        "created_at": r.get("created_at"),
    }


def alert_to_dict(a: Any) -> dict:
    return {
        "id": str(a.get("id")),
        "enterprise_id": str(a.get("enterprise_id")),
        "level": a.get("level"),
        "alert_type": a.get("alert_type"),
        "message": a.get("message"),
        "due_at": a.get("due_at"),
        "acked": a.get("acked"),
        "created_at": a.get("created_at"),
    }
